from colorama import Fore, init
import os

init(autoreset=True)

file_name = "Phone Number List.txt"
from colorama import Fore, init
from openpyxl import Workbook, load_workbook
import os

init(autoreset=True)

file_name = "MapyCRM.xlsx"

# Create Excel file if it doesn't exist
if not os.path.exists(file_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.append(["Name", "Phone"])  # headers
    wb.save(file_name)

print("Welcome To MapyCRM")
print("Type 'q' anytime to quit\n")

while True:
    name = input("Client name: ").strip()
    if name.lower() == "q":
        print("Goodbye 👋")
        break

    number = input("Phone number (+212): ").strip()
    if number.lower() == "q":
        print("Goodbye 👋")
        break

    phone = f"+212 {number}"

    wb = load_workbook(file_name)
    ws = wb.active

    # Check if phone already exists
    phones = [row[1].value for row in ws.iter_rows(min_row=2)]

    if phone in phones:
        print(Fore.RED + "Phone number already exists ❌")
    else:
        ws.append([name, phone])
        wb.save(file_name)
        print(Fore.GREEN + f"{name} | {phone} Added Successfully ✅")
